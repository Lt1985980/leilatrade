import argparse
import asyncio
import functools
import gc
import json
import logging
import os
import re
import sqlite3
import sys
import time

from collections import defaultdict
from contextlib import contextmanager
from datetime import datetime, timedelta, timezone
from logging.handlers import RotatingFileHandler
from typing import Any, Dict, List, Optional, Tuple, Union

import aiohttp
import ccxt
import numpy as np
import pandas as pd
import psutil
import structlog
import talib
from aiohttp import ClientTimeout, TCPConnector, web
from cachetools import TTLCache
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from prometheus_client import Counter, Gauge, Histogram, generate_latest

# --------- LOAD ENV ----------
load_dotenv() 

# ========== CONFIG ==========
MIN_SIGNAL_CONFIDENCE = float(os.getenv("MIN_SIGNAL_CONFIDENCE", "49"))
STRONG_SIGNAL_THRESHOLD = 69
WEAK_SIGNAL_MAX = 57
RISK_PER_TRADE = float(os.getenv("RISK_PER_TRADE", "0.02"))
MAX_POSITION_SIZE = float(os.getenv("MAX_POSITION_SIZE", "0.1"))

# ========== تنظیمات فیلترهای پیشرفته ==========
FILTER_CONFIG = {
    'confidence_filter': {
        'min_confidence': MIN_SIGNAL_CONFIDENCE,
        'strong_threshold': STRONG_SIGNAL_THRESHOLD
    },
    'risk_filter': {
        'max_risk_per_trade': RISK_PER_TRADE,
        'max_position_size': MAX_POSITION_SIZE
    },
    'volume_filter': {
        'min_volume_btc': 0.1,
        'min_volume_ratio': 0.8
    },
    'timeframe_filter': {
        'required_confirmations': 2,
        'priority_timeframes': ['1h', '4h', '30m']
    }
}

# ========== FEATURE FLAGS ==========
FEATURE_FLAGS = {
    'advanced_ml': os.getenv("ENABLE_ADVANCED_ML", "true").lower() == "true",
    'news_analysis': os.getenv("ENABLE_NEWS_ANALYSIS", "false").lower() == "true",
    'telegram_alerts': os.getenv("ENABLE_TELEGRAM", "true").lower() == "true",
    'excel_reports': os.getenv("ENABLE_EXCEL", "true").lower() == "true",
    'health_server': os.getenv("ENABLE_HEALTH_SERVER", "true").lower() == "true",
}

# ========== تنظیمات امنیتی و شبکه ==========
CIRCUIT_BREAKER_THRESHOLD = int(os.getenv("CIRCUIT_BREAKER_THRESHOLD", "5"))
MAX_RETRIES = int(os.getenv("MAX_RETRIES", "3"))
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "20"))

# ========== تنظیمات هوش مصنوعی ==========
USE_ML_PREDICTIONS = os.getenv("USE_ML_PREDICTIONS", "true").lower() == "true"

# ========== تنظیمات مالی ==========
INITIAL_BALANCE = float(os.getenv("INITIAL_BALANCE", "10000"))

# ========== تنظیمات دیتابیس و خروجی ==========
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)
LOG_FILE = os.path.join(OUTPUT_DIR, "bot.log")
DB_PATH = os.path.join(OUTPUT_DIR, "signals.db")

# ========== تنظیمات اجرا ==========
RUN_INTERVAL = int(os.getenv("RUN_INTERVAL", "3600"))
TIMEFRAMES = tuple(os.getenv("TIMEFRAMES", "15m,30m,1h,4h,1d").split(","))
SYMBOLS_BASE = tuple(os.getenv("SYMBOLS_BASE", "BTC,ETH,SOL,ADA,XRP,DOT,BNB").split(","))
SYMBOLS = [f"{b}/USDT" for b in SYMBOLS_BASE]

# ========== تنظیمات APIها ==========
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")
TELEGRAM_ENABLED = bool(TELEGRAM_TOKEN and TELEGRAM_CHAT_ID)

COINGECKO_API_KEY = os.getenv("COINGECKO_API_KEY", "")
NEWSAPI_KEY = os.getenv("NEWSAPI_KEY", "")
CRYPTOPANIC_API_KEY = os.getenv("CRYPTOPANIC_API_KEY", "")
COINMARKETCAP_API_KEY = os.getenv("COINMARKETCAP_API_KEY", "")

class Config:
    """کلاس کانفیگ برای دسترسی متمرکز به تنظیمات"""
    
    def __init__(self):
        # سیگنال و ریسک
        self.MIN_SIGNAL_CONFIDENCE = MIN_SIGNAL_CONFIDENCE
        self.STRONG_SIGNAL_THRESHOLD = STRONG_SIGNAL_THRESHOLD
        self.WEAK_SIGNAL_MAX = WEAK_SIGNAL_MAX
        self.RISK_PER_TRADE = RISK_PER_TRADE
        self.MAX_POSITION_SIZE = MAX_POSITION_SIZE
        
        # فیلترها
        self.FILTER_CONFIG = FILTER_CONFIG
        
        # امنیت و شبکه
        self.CIRCUIT_BREAKER_THRESHOLD = CIRCUIT_BREAKER_THRESHOLD
        self.MAX_RETRIES = MAX_RETRIES
        self.REQUEST_TIMEOUT = REQUEST_TIMEOUT
        
        # هوش مصنوعی
        self.USE_ML_PREDICTIONS = USE_ML_PREDICTIONS
        
        # مالی
        self.INITIAL_BALANCE = INITIAL_BALANCE
        
        # دیتابیس و خروجی
        self.OUTPUT_DIR = OUTPUT_DIR
        self.LOG_FILE = LOG_FILE
        self.DB_PATH = DB_PATH
        
        # اجرا
        self.RUN_INTERVAL = RUN_INTERVAL
        self.TIMEFRAMES = TIMEFRAMES
        self.SYMBOLS_BASE = SYMBOLS_BASE
        self.SYMBOLS = SYMBOLS
        
        # APIها
        self.TELEGRAM_TOKEN = TELEGRAM_TOKEN
        self.TELEGRAM_CHAT_ID = TELEGRAM_CHAT_ID
        self.TELEGRAM_ENABLED = TELEGRAM_ENABLED
        self.COINGECKO_API_KEY = COINGECKO_API_KEY
        self.NEWSAPI_KEY = NEWSAPI_KEY
        self.CRYPTOPANIC_API_KEY = CRYPTOPANIC_API_KEY
        self.COINMARKETCAP_API_KEY = COINMARKETCAP_API_KEY
        
        # Feature Flags
        self.FEATURE_FLAGS = FEATURE_FLAGS

# ایجاد نمونه کانفیگ
config = Config()

# ========== VALIDATION FUNCTIONS ==========
def validate_symbol(symbol: str) -> bool:
    """
    اعتبارسنجی فرمت سیمبل
    
    Args:
        symbol (str): سیمبل به فرمت BASE/QUOTE
        
    Returns:
        bool: True اگر فرمت معتبر باشد
        
    Example:
        >>> validate_symbol("BTC/USDT")
        True
        >>> validate_symbol("invalid")
        False
    """
    return re.match(r"^[A-Z]+/[A-Z]+$", symbol) is not None

# ========== RISK MANAGER FUNCTIONS ==========
def calculate_position_size(signal, account_balance):
    """
    محاسبه اندازه پوزیشن با مدیریت ریسک پیشرفته
    
    Args:
        signal (dict): دیکشنری سیگنال شامل:
            - confidence (float): اطمینان سیگنال (0-100)
            - signal (str): نوع سیگنال (BUY/SELL/HOLD)
            - stop_loss (float): قیمت حد ضرر
            - current_price (float): قیمت فعلی
        account_balance (float): موجودی کل حساب
        
    Returns:
        float: اندازه پوزیشن بر اساس واحد ارز پایه
        
    Raises:
        ValueError: اگر پارامترهای ضروری missing باشند
        
    Example:
        >>> signal = {'confidence': 75, 'stop_loss': 49000, 'current_price': 50000}
        >>> size = calculate_position_size(signal, 10000)
        >>> print(f"Position size: {size:.4f}")
    """
    confidence = signal.get('confidence', 0)
    signal_type = signal.get('signal', 'HOLD')
    
    # پایه ریسک بر اساس تنظیمات
    base_risk = RISK_PER_TRADE
    
    # تنظیم ریسک بر اساس اطمینان سیگنال
    risk_multiplier = 1.0
    if confidence >= STRONG_SIGNAL_THRESHOLD:
        risk_multiplier = 1.2  # 20% افزایش برای سیگنال‌های قوی
    elif confidence >= 70:
        risk_multiplier = 1.0  # ریسک معمولی
    elif confidence >= MIN_SIGNAL_CONFIDENCE:
        risk_multiplier = 0.7  # 30% کاهش برای سیگنال‌های حاشیه‌ای
    
    # محاسبه ریسک نهایی
    final_risk = base_risk * risk_multiplier
    final_risk = min(final_risk, MAX_POSITION_SIZE)
    
    position_size = account_balance * final_risk
    stop_loss = signal.get('stop_loss', 0)
    current_price = signal.get('current_price', 0)
    
    # محاسبه تعداد واحدها
    if stop_loss and current_price:
        risk_per_unit = abs(current_price - stop_loss)
        if risk_per_unit > 0:
            units = (account_balance * final_risk) / risk_per_unit
            return min(units, position_size / current_price)
    
    return position_size / current_price

def validate_risk_parameters(signal):
    """اعتبارسنجی پارامترهای ریسک"""
    sl = signal.get('stop_loss', 0)
    tp = signal.get('take_profit', 0)
    price = signal.get('current_price', 0)
    
    if not all([sl, tp, price]):
        return False, "پارامترهای ریسک ناقص"
    
    # بررسی نسبت ریسک به پاداش
    risk = abs(price - sl)
    reward = abs(tp - price)
    risk_reward_ratio = reward / risk if risk > 0 else 0
    
    if risk_reward_ratio < 1.2:  # حداقل نسبت 1:1.2
        return False, f"نسبت Risk/Reward نامناسب: {risk_reward_ratio:.2f}"
    
    # بررسی فاصله استاپ لاس
    stop_loss_percent = abs(price - sl) / price * 100
    if stop_loss_percent > 10:  # حداکثر استاپ لاس 10%
        return False, f"استاپ لاس بسیار بزرگ: {stop_loss_percent:.1f}%"
    
    return True, "پارامترهای ریسک معتبر"

# ========== SIGNAL FILTER FUNCTIONS ==========
def calculate_volume_score(signal):
    """امتیازدهی بر اساس حجم معاملات"""
    volume = signal.get('volume', 0)
    symbol = signal.get('symbol', '')
    
    volume_thresholds = {
        'BTC/USDT': 100, 'ETH/USDT': 1000, 'SOL/USDT': 5000,
        'ADA/USDT': 500000, 'XRP/USDT': 300000, 'DOT/USDT': 30000, 'BNB/USDT': 500
    }
    
    min_volume = volume_thresholds.get(symbol, 0)
    return 1.0 if volume >= min_volume else volume / min_volume

def calculate_timeframe_score(signal):
    """امتیازدهی تایم‌فریم"""
    timeframe = signal.get('timeframe', '')
    priority_tf = ['1h', '4h', '30m']  # تایم‌فریم‌های اولویت‌دار
    return 1.0 if timeframe in priority_tf else 0.7

def calculate_signal_score(signal):
    """محاسبه امتیاز ترکیبی برای سیگنال"""
    base_score = signal.get('confidence', 0) / 100.0
    technical_score = signal.get('score', 0)
    
    # فاکتورهای امتیازدهی
    factors = {
        'confidence_factor': base_score * 0.4,
        'technical_score_factor': abs(technical_score) * 0.3,
        'volume_factor': calculate_volume_score(signal) * 0.2,
        'timeframe_factor': calculate_timeframe_score(signal) * 0.1
    }
    
    total_score = sum(factors.values())
    return min(total_score * 100, 100)

def apply_final_filters(signals, account_balance=1000):
    """اعمال فیلترهای نهایی بر سیگنال‌ها"""
    filtered_signals = []
    
    for signal in signals:
        # فیلتر 1: اطمینان حداقل
        if signal.get('confidence', 0) < MIN_SIGNAL_CONFIDENCE:
            continue
            
        # فیلتر 2: امتیاز ترکیبی
        composite_score = calculate_signal_score(signal)
        if composite_score < 57:  # حداقل امتیاز ترکیبی
            continue
            
        # فیلتر 3: اعتبارسنجی ریسک
        is_risk_valid, risk_message = validate_risk_parameters(signal)
        if not is_risk_valid:
            logger.warning(f"❌ رد سیگنال به دلیل ریسک: {signal.get('symbol', '')} - {risk_message}")
            continue
            
        # فیلتر 4: محاسبه اندازه پوزیشن
        position_size = calculate_position_size(signal, account_balance)
        if position_size <= 0:
            continue
            
        # افزودن اطلاعات تکمیلی
        signal['composite_score'] = composite_score
        signal['position_size'] = position_size
        signal['risk_percentage'] = RISK_PER_TRADE * 100
        
        filtered_signals.append(signal)
    
    # مرتب‌سازی بر اساس امتیاز ترکیبی
    filtered_signals.sort(key=lambda x: x.get('composite_score', 0), reverse=True)
    
    return filtered_signals

def log_signal_analysis(signals_before, signals_after):
    """گزارش تحلیل کیفیت سیگنال‌ها"""
    logger.info("🎯 تحلیل نهایی سیگنال‌ها")
    logger.info(f"📊 قبل از فیلتر: {len(signals_before)} سیگنال")
    logger.info(f"📊 بعد از فیلتر: {len(signals_after)} سیگنال")
    
    if signals_after:
        avg_confidence = sum(s.get('confidence', 0) for s in signals_after) / len(signals_after)
        avg_score = sum(s.get('composite_score', 0) for s in signals_after) / len(signals_after)
        logger.info(f"📈 میانگین اطمینان: {avg_confidence:.1f}%")
        logger.info(f"📈 میانگین امتیاز ترکیبی: {avg_score:.1f}%")
        
        # نمایش سیگنال‌های برتر
        logger.info("🏆 سیگنال‌های برتر:")
        for i, signal in enumerate(signals_after[:3]):
            logger.info(f"   {i+1}. {signal.get('symbol', '')} {signal.get('timeframe', '')} | "
                       f"سیگنال: {signal.get('signal', '')} | "
                       f"اطمینان: {signal.get('confidence', 0)}% | "
                       f"امتیاز: {signal.get('composite_score', 0):.1f}%")

# --------- STRUCTURED LOGGING ----------
def setup_structured_logging():
    try:
        structlog.configure(
            processors=[
                structlog.stdlib.add_log_level,
                structlog.processors.TimeStamper(fmt="iso"),
                structlog.processors.JSONRenderer(),
            ],
            wrapper_class=structlog.stdlib.BoundLogger,
            logger_factory=structlog.stdlib.LoggerFactory(),
            cache_logger_on_first_use=True,
        )
        return structlog.get_logger()
    except Exception:
        return None

structured_logger = setup_structured_logging()

logger = logging.getLogger("crypto_analyzer")
logger.setLevel(logging.INFO)
logger.propagate = False
fmt = logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s", "%Y-%m-%d %H:%M:%S")
fh = RotatingFileHandler(config.LOG_FILE, maxBytes=5_000_000, backupCount=3, encoding="utf-8")
fh.setFormatter(fmt)
ch = logging.StreamHandler(sys.stdout)
ch.setFormatter(fmt)
logger.handlers.clear()
logger.addHandler(fh)
logger.addHandler(ch)
logger.info("پوشه خروجی آماده است: %s", config.OUTPUT_DIR)

# --------- PROMETHEUS METRICS ----------
try:
    REQUESTS_TOTAL = Counter("requests_total", "Total HTTP requests", ["method", "endpoint", "status"])
    REQUEST_DURATION = Histogram("request_duration_seconds", "HTTP request duration seconds")
    PRICE_SOURCE_SUCCESS = Gauge("price_source_success_rate", "Success rate per price source", ["source"])
    OHLCV_SOURCE_SUCCESS = Gauge("ohlcv_source_success", "OHLCV source success flag", ["source"])
    OHLCV_FETCH_FAILURES = Counter("ohlcv_fetch_failures_total", "OHLCV fetch failures", ["symbol"])
    ACTIVE_SIGNALS = Gauge("active_signals", "Number of active signals")
except Exception as e:
    logger.debug(f"خطای اولیه‌سازی متریک‌ها: {e}")
    REQUESTS_TOTAL = None
    REQUEST_DURATION = None
    PRICE_SOURCE_SUCCESS = None
    OHLCV_SOURCE_SUCCESS = None
    OHLCV_FETCH_FAILURES = None
    ACTIVE_SIGNALS = None

# --------- CACHE INITIALIZATION ----------
WEIGHTED_PRICE_CACHE = TTLCache(maxsize=100, ttl=120)
NEWS_CACHE = TTLCache(maxsize=100, ttl=600)

# --------- UTILITY ----------
def safe_float(value, default=0.0):
    try:
        return float(value) if value is not None else default
    except (ValueError, TypeError):
        return default

def safe_get(series, index=-1, default=0):
    try:
        if series is None:
            return default
        if isinstance(series, (list, tuple, np.ndarray)):
            if len(series) == 0:
                return default
            try:
                return series[index]
            except Exception:
                return default
        if hasattr(series, "empty") and series.empty:
            return default
        if index < 0:
            idx = len(series) + index
        else:
            idx = index
        if idx < 0 or idx >= len(series):
            return default
        value = series.iloc[idx] if hasattr(series, "iloc") else series[idx]
        return value if not pd.isna(value) else default
    except Exception:
        return default

def fmt_num(value, digits=6, default="-"):
    try:
        v = float(value)
        return f"{v:.{digits}f}"
    except Exception:
        return default

# --------- CIRCUIT BREAKER ----------
class CircuitBreaker:
    def __init__(self, failure_threshold=5, reset_timeout=60):
        self.failure_threshold = failure_threshold
        self.reset_timeout = reset_timeout
        self.failure_count = 0
        self.last_failure_time = None
        self.state = "CLOSED"  # CLOSED, OPEN, HALF_OPEN

    async def call(self, coro):
        if self.state == "OPEN":
            if self.last_failure_time and (time.time() - self.last_failure_time > self.reset_timeout):
                self.state = "HALF_OPEN"
                logger.debug("مدار به HALF_OPEN رفت")
            else:
                raise Exception("Circuit breaker باز است - سرویس موقتاً در دسترس نیست")
        try:
            result = await coro
            self._on_success()
            return result
        except Exception:
            self._on_failure()
            raise

    def _on_success(self):
        self.failure_count = 0
        self.last_failure_time = None
        self.state = "CLOSED"

    def _on_failure(self):
        self.failure_count += 1
        self.last_failure_time = time.time()
        if self.failure_count >= self.failure_threshold:
            self.state = "OPEN"
            logger.warning(f"مدار حفاظتی پس از {self.failure_count} خطا باز شد")

    def get_status(self):
        return {"state": self.state, "failure_count": self.failure_count, "last_failure_time": self.last_failure_time}

http_cb = CircuitBreaker(failure_threshold=config.CIRCUIT_BREAKER_THRESHOLD, reset_timeout=60)

# --------- SMART CACHE ----------
class SmartCache:
    def __init__(self, maxsize=200, ttl=300):
        self.cache = TTLCache(maxsize=maxsize, ttl=ttl)
        self.success_count = defaultdict(int)
        self.total_requests = defaultdict(int)
        self.hits = 0
        self.misses = 0
        self._lock = asyncio.Lock()

    def get_success_rate(self, key):
        if self.total_requests[key] == 0:
            return 0.0
        rate = self.success_count[key] / self.total_requests[key]
        try:
            if PRICE_SOURCE_SUCCESS:
                PRICE_SOURCE_SUCCESS.labels(source=key).set(self.get_success_rate(key))
        except Exception:
            pass
        return rate

    def record_success(self, key):
        self.success_count[key] += 1
        self.total_requests[key] += 1

    def record_failure(self, key):
        self.total_requests[key] += 1
        try:
            if PRICE_SOURCE_SUCCESS:
                PRICE_SOURCE_SUCCESS.labels(source=key).set(self.get_success_rate(key))
        except Exception:
            pass

    def get_hit_rate(self):
        total = self.hits + self.misses
        return self.hits / total if total > 0 else 0.0

    def __contains__(self, key):
        return key in self.cache

    def __getitem__(self, key):
        if key in self.cache:
            self.hits += 1
            return self.cache[key]
        self.misses += 1
        raise KeyError(key)

    def __setitem__(self, key, value):
        self.cache[key] = value

    def __delitem__(self, key):
        del self.cache[key]

    def get(self, key, default=None):
        try:
            return self[key]
        except KeyError:
            return default

    def clear(self):
        self.cache.clear()
        self.hits = 0
        self.misses = 0

    def info(self):
        return {
            "cache_size": len(self.cache),
            "max_size": self.cache.maxsize,
            "hits": self.hits,
            "misses": self.misses,
            "hit_rate": self.get_hit_rate(),
            "success_rates": {k: self.get_success_rate(k) for k in self.total_requests},
        }

    async def get_or_set(self, key, coroutine, *args, **kwargs):
        async with self._lock:
            cached_value = self.get(key)
            if cached_value is not None:
                return cached_value
            result = await coroutine(*args, **kwargs)
            if result is not None:
                self[key] = result
            return result

def async_cached(cache: Union[SmartCache, TTLCache]):
    """
    Decorator برای کش نتایج توابع async.
    نکته: session را از کلید کش حذف می‌کنیم تا سریال‌سازی امن شود.
    """
    def decorator(func):
        @functools.wraps(func)
        async def wrapper(*args, **kwargs):
            try:
                safe_args = [a for a in args if not isinstance(a, aiohttp.ClientSession)]
                key = json.dumps({"fn": func.__name__, "args": safe_args, "kwargs": kwargs}, default=str, sort_keys=True)
            except Exception:
                key = func.__name__ + str(args) + str(kwargs)
            try:
                if key in cache:
                    return cache[key]
            except Exception:
                pass
            result = await func(*args, **kwargs)
            try:
                cache[key] = result
            except Exception:
                pass
            return result
        return wrapper
    return decorator

# --------- CACHE INSTANCES ----------
PRICE_CACHE = SmartCache(maxsize=300, ttl=300)

# --------- DATABASE ----------
@contextmanager
def get_db_connection():
    conn = sqlite3.connect(config.DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def init_db():
    with get_db_connection() as conn:
        cur = conn.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS signals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT,
                symbol TEXT NOT NULL,
                timeframe TEXT NOT NULL,
                signal TEXT,
                score REAL,
                confidence REAL,
                price REAL,
                sl REAL,
                tp REAL,
                news_score REAL,
                price_rel REAL,
                news_rel REAL,
                ml_agreement INTEGER,
                ml_confidence REAL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(symbol, timeframe, ts)
            )
            """
        )
        logger.info("دیتابیس با موفقیت آماده شد")

def update_database_schema():
    with get_db_connection() as conn:
        cur = conn.cursor()
        try:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='signals'")
            if cur.fetchone() is None:
                init_db()
                return
            cur.execute("PRAGMA table_info(signals)")
            columns = [col[1] for col in cur.fetchall()]
            required_columns = {
                "sl": "REAL", "tp": "REAL", "ml_agreement": "INTEGER",
                "ml_confidence": "REAL", "news_score": "REAL",
                "price_rel": "REAL", "news_rel": "REAL", "price": "REAL"
            }
            changed = False
            for column_name, column_type in required_columns.items():
                if column_name not in columns:
                    try:
                        cur.execute(f"ALTER TABLE signals ADD COLUMN {column_name} {column_type}")
                        logger.info(f"ستون اضافه شد: {column_name}")
                        changed = True
                    except sqlite3.Error as column_error:
                        logger.error(f"خطا در افزودن ستون {column_name}: {column_error}")
                        continue
            if changed:
                logger.info("اسکیما دیتابیس با موفقیت به‌روز شد")
            else:
                logger.info("اسکیما دیتابیس از قبل به‌روز است")
        except Exception as e:
            logger.error(f"خطا در به‌روزرسانی اسکیما: {e}")

def log_signal_to_db(sig: Dict[str, Any]):
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute(
                """
                INSERT OR IGNORE INTO signals (
                    ts, symbol, timeframe, signal, score, confidence,
                    price, sl, tp, news_score, price_rel, news_rel,
                    ml_agreement, ml_confidence
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    sig.get("ts"),
                    sig.get("symbol"),
                    sig.get("timeframe"),
                    sig.get("signal"),
                    safe_float(sig.get("score", 0.0)),
                    safe_float(sig.get("confidence", 0)),
                    safe_float(sig.get("entry_price", sig.get("current_price", 0))),
                    safe_float(sig.get("stop_loss", sig.get("sl", 0))),
                    safe_float(sig.get("take_profit", sig.get("tp", 0))),
                    safe_float(sig.get("news_score", 0)),
                    safe_float(sig.get("price_reliability", sig.get("price_rel", 0))),
                    safe_float(sig.get("news_reliability", sig.get("news_rel", 0))),
                    1 if sig.get("ml_agreement") else 0,
                    safe_float(sig.get("ml_confidence", 0.0)),
                ),
            )
            logger.info(f"سیگنال ذخیره شد: {sig.get('symbol')} {sig.get('timeframe')}")
    except Exception as e:
        logger.error(f"خطا در ذخیره سیگنال: {e}")

# --------- COMMON HTTP ----------
USER_AGENT = "LeilaTrader/2.0"
HEADERS_DEFAULT = {"User-Agent": USER_AGENT, "Accept": "application/json"}

async def http_get(
    session: aiohttp.ClientSession,
    url: str,
    params: Optional[Dict] = None,
    headers: Optional[Dict] = None,
    retries: int = MAX_RETRIES,
    backoff: float = 0.9,
    timeout: int = REQUEST_TIMEOUT,
) -> Optional[Dict]:
    async def _do():
        merged_headers = {**HEADERS_DEFAULT, **(headers or {})}
        for i in range(retries):
            start_time = time.time()
            try:
                async with session.get(
                    url,
                    params=params,
                    headers=merged_headers,
                    timeout=ClientTimeout(total=timeout),
                ) as resp:
                    text = await resp.text()
                    try:
                        if REQUESTS_TOTAL:
                            REQUESTS_TOTAL.labels(method="GET", endpoint=url, status=resp.status).inc()
                    except Exception:
                        pass
                    if resp.status == 200:
                        try:
                            if REQUEST_DURATION:
                                REQUEST_DURATION.observe(time.time() - start_time)
                        except Exception:
                            pass
                        try:
                            return await resp.json()
                        except Exception:
                            logger.debug(f"خطا در پارس JSON برای {url}: {text[:200]}")
                            return None
                    else:
                        if resp.status in (429, 500, 502, 503, 504):
                            logger.debug(f"پاسخ غیر 200 {resp.status} از {url} - تلاش مجدد")
                        else:
                            logger.debug(f"پاسخ {resp.status} از {url}: {text[:200]}")
            except Exception as e:
                logger.debug(f"خطای GET برای {url}: {e}")
            if i < retries - 1:
                await asyncio.sleep(backoff * (2**i))
        return None
    # اعمال CircuitBreaker روی فراخوانی داخلی
    return await http_cb.call(_do())

# --------- RISK MANAGER ----------
class RiskManager:
    def __init__(self, max_position_size: float = config.MAX_POSITION_SIZE):
        self.max_position_size = max_position_size

    def calculate_position_size(
        self,
        account_balance: float,
        entry_price: float,
        stop_loss: float,
        confidence: float,
    ) -> float:
        risk_per_trade = account_balance * config.RISK_PER_TRADE
        price_risk = abs(entry_price - stop_loss)
        if price_risk <= 0:
            return 0.0
        position_size = risk_per_trade / price_risk
        adjusted_size = position_size * max(0.0, min(1.0, confidence / 100.0))
        return min(adjusted_size, account_balance * self.max_position_size)

# --------- PRICE FETCHERS ----------
@async_cached(PRICE_CACHE)
async def fetch_price_mexc(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:
    try:
        market = symbol.replace("/", "").upper()
        url = f"https://api.mexc.com/api/v3/ticker/price?symbol={market}"
        data = await http_get(session, url, timeout=10)
        if data and "price" in data:
            PRICE_CACHE.record_success("mexc")
            return float(data["price"])
    except Exception as e:
        logger.debug(f"خطا در MEXC برای {symbol}: {e}")
    PRICE_CACHE.record_failure("mexc")
    return None

@async_cached(PRICE_CACHE)
async def fetch_price_toobit(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:
    try:
        market = symbol.replace("/", "")
        url = f"https://api.toobit.com/v5/market/tickers?category=spot&symbol={market}"
        data = await http_get(session, url, timeout=10)
        if (data and isinstance(data, dict) and "result" in data and 
            isinstance(data["result"], dict) and "list" in data["result"] and 
            isinstance(data["result"]["list"], list) and len(data["result"]["list"]) > 0 and 
            "lastPrice" in data["result"]["list"][0]):
            PRICE_CACHE.record_success("toobit")
            return float(data["result"]["list"][0]["lastPrice"])
    except Exception as e:
        logger.debug(f"خطا در Toobit برای {symbol}: {e}")
    PRICE_CACHE.record_failure("toobit")
    return None

@async_cached(PRICE_CACHE)
async def fetch_price_coingecko(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:
    try:
        coin_map = {
            "BTC/USDT": "bitcoin", "ETH/USDT": "ethereum", "BNB/USDT": "binancecoin",
            "ADA/USDT": "cardano", "SOL/USDT": "solana", "XRP/USDT": "ripple",
            "DOT/USDT": "polkadot",
        }
        coin_id = coin_map.get(symbol)
        if not coin_id:
            return None
        url = "https://api.coingecko.com/api/v3/simple/price"
        params = {"ids": coin_id, "vs_currencies": "usd"}
        headers = {"x-cg-demo-api-key": config.COINGECKO_API_KEY} if config.COINGECKO_API_KEY else {}
        data = await http_get(session, url, params=params, headers=headers, timeout=10)
        if data and coin_id in data and "usd" in data[coin_id]:
            PRICE_CACHE.record_success("coingecko")
            return float(data[coin_id]["usd"])
    except Exception as e:
        logger.debug(f"خطا در CoinGecko برای {symbol}: {e}")
    PRICE_CACHE.record_failure("coingecko")
    return None

@async_cached(PRICE_CACHE)
async def fetch_price_coinmarketcap(session: aiohttp.ClientSession, symbol: str) -> Optional[float]:
    try:
        base = symbol.split("/")[0]
        url = "https://pro-api.coinmarketcap.com/v1/cryptocurrency/quotes/latest"
        headers = {"X-CMC_PRO_API_KEY": config.COINMARKETCAP_API_KEY} if config.COINMARKETCAP_API_KEY else {}
        params = {"symbol": base, "convert": "USD"}
        data = await http_get(session, url, params=params, headers=headers, timeout=10)
        if data and "data" in data and base in data["data"]:
            PRICE_CACHE.record_success("coinmarketcap")
            return float(data["data"][base]["quote"]["USD"]["price"])
    except Exception as e:
        logger.debug(f"خطا در CoinMarketCap برای {symbol}: {e}")
    PRICE_CACHE.record_failure("coinmarketcap")
    return None

PRICE_SOURCE_WEIGHTS = {
    "mexc": 0.25, "toobit": 0.25, "coinmarketcap": 0.25, "coingecko": 0.25,
}

@async_cached(WEIGHTED_PRICE_CACHE)
async def fetch_price_weighted(session: aiohttp.ClientSession, symbol: str):
    # اعتبارسنجی سیمبل
    if not validate_symbol(symbol):
        logger.error(f"فرمت سیمبل نامعتبر: {symbol}")
        return None, 0.0, {}
        
    try:
        tasks = {
            'mexc': fetch_price_mexc(session, symbol),
            'toobit': fetch_price_toobit(session, symbol),
            'coingecko': fetch_price_coingecko(session, symbol),
            'coinmarketcap': fetch_price_coinmarketcap(session, symbol),
        }
        
        results = await asyncio.gather(*tasks.values(), return_exceptions=True)
        
        source_prices: Dict[str, Optional[float]] = {}
        active: Dict[str, Tuple[float, float]] = {}
        
        for (name, _), result in zip(tasks.items(), results):
            if isinstance(result, Exception):
                logger.debug(f"خطا در منبع {name} برای {symbol}: {result}")
                price = None
            else:
                price = result
            
            source_prices[name] = price
            
            if price is not None and isinstance(price, (int, float)):
                base_w = PRICE_SOURCE_WEIGHTS.get(name, 0.15)
                sr = PRICE_CACHE.get_success_rate(name)
                dyn_w = max(0.1, min(0.4, base_w * (0.8 + 0.4 * sr)))
                active[name] = (float(price), dyn_w)
            else:
                logger.debug(f"منبع قیمت ناموفق: {name} برای {symbol}")

        if not active:
            logger.warning(f"هیچ منبع قیمتی فعال برای {symbol} نبود؛ تلاش fallback")
            retry_cg = await fetch_price_coingecko(session, symbol)
            retry_cmc = await fetch_price_coinmarketcap(session, symbol)
            candidates = [p for p in [retry_cg, retry_cmc] if p is not None and isinstance(p, (int, float))]
            
            if candidates:
                final_price = float(np.mean(candidates))
                reliability = 0.25
                logger.info(f"Fallback قیمت برای {symbol}: {fmt_num(final_price)}")
                return final_price, reliability, source_prices
            else:
                logger.error(f"همه منابع قیمت برای {symbol} ناموفق بودند")
                return None, 0.0, source_prices

        if len(active) == 1:
            (name, (price, _)) = list(active.items())[0]
            logger.info(f"یک منبع فعال برای {symbol}: {name} → {fmt_num(price)}")
            return price, 0.35, source_prices

        total_w = sum(w for _, w in active.values())
        weighted_price = sum(p * (w / total_w) for p, w in active.values())
        reliability = len(active) / len(tasks)
        
        logger.info(f"قیمت وزن‌دهی‌شده {symbol} = {fmt_num(weighted_price)} | اتکا={reliability*100:.1f}% | منابع={list(active.keys())}")
        return float(weighted_price), float(reliability), source_prices
        
    except Exception as e:
        logger.error(f"خطا در وزن‌دهی قیمت برای {symbol}: {e}")
        return None, 0.0, {}

async def fetch_price_best(session: aiohttp.ClientSession, symbol: str) -> Tuple[Optional[float], float]:
    result = await fetch_price_weighted(session, symbol)
    if result:
        price, reliability, _ = result
        return price, reliability
    return None, 0.0

# --------- NEWS FETCHERS ----------
def simple_sentiment(text: str) -> float:
    txt = (text or "").lower()
    pos = ['rise', 'bull', 'gain', 'positive', 'up', 'surge', 'pump', 'rally']
    neg = ['fall', 'bear', 'loss', 'negative', 'down', 'dump', 'plunge']
    score = 0.0
    score += sum(1 for w in pos if w in txt) * 0.2
    score -= sum(1 for w in neg if w in txt) * 0.2
    return max(-1.0, min(1.0, score))

def recency_boost(published_at: Optional[str]) -> float:
    try:
        if not published_at:
            return 0.5
        pub_time = datetime.fromisoformat(published_at.replace('Z', '+00:00'))
        now = datetime.now(pub_time.tzinfo)
        hours_ago = (now - pub_time).total_seconds() / 3600
        if hours_ago < 1:
            return 1.0
        elif hours_ago < 6:
            return 0.8
        elif hours_ago < 24:
            return 0.5
        else:
            return 0.2
    except Exception:
        return 0.5

@async_cached(NEWS_CACHE)
async def fetch_news_total(session: aiohttp.ClientSession, symbol: str) -> Tuple[int, float, float]:
    try:
        base = symbol.split("/")[0]
        if config.NEWSAPI_KEY:
            return 5, 0.5 if not config.CRYPTOPANIC_API_KEY else 0.7, 0.1
        elif config.CRYPTOPANIC_API_KEY:
            return 3, 0.4, 0.05
        else:
            return 0, 0.0, 0.0
    except Exception as e:
        logger.debug(f"خطای news برای {symbol}: {e}")
        return 0, 0.0, 0.0

# --------- OHLCV FETCHERS ----------
async def fetch_mexc_klines(session: aiohttp.ClientSession, symbol: str, timeframe: str, limit: int) -> pd.DataFrame:
    try:
        market = symbol.replace("/", "").upper()
        interval_map = {"15m": "15m", "30m": "30m", "1h": "60m", "4h": "4h", "1d": "1d"}
        interval = interval_map.get(timeframe, "1h")
        url = "https://api.mexc.com/api/v3/klines"
        params = {"symbol": market, "interval": interval, "limit": limit}
        data = await http_get(session, url, params=params)
        if data and isinstance(data, list) and len(data) > 0 and len(data[0]) >= 6:
            df = pd.DataFrame(data)
            df = df.iloc[:, :6]
            df.columns = ["timestamp", "open", "high", "low", "close", "volume"]
            df["timestamp"] = pd.to_datetime(df["timestamp"], unit="ms", utc=True)
            df = df.set_index("timestamp")
            for col in ["open", "high", "low", "close", "volume"]:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            df = df[["open", "high", "low", "close", "volume"]].dropna()
            try:
                if OHLCV_SOURCE_SUCCESS:
                    OHLCV_SOURCE_SUCCESS.labels(source="mexc").set(1)
            except Exception:
                pass
            return df
    except Exception as e:
        logger.debug(f"خطا در کندل‌های MEXC برای {symbol}: {e}")
    return pd.DataFrame()

async def fetch_toobit_klines(session: aiohttp.ClientSession, symbol: str, timeframe: str, limit: int) -> pd.DataFrame:
    try:
        market = symbol.replace("/", "")
        interval_map = {"15m": "15", "30m": "30", "1h": "60", "4h": "240", "1d": "D"}
        interval = interval_map.get(timeframe, "60")
        url = "https://api.toobit.com/v5/market/kline"
        params = {"category": "spot", "symbol": market, "interval": interval, "limit": limit}
        data = await http_get(session, url, params=params)
        if data and "result" in data and "list" in data["result"]:
            klines = data["result"]["list"]
            if klines:
                klines.reverse()
                df = pd.DataFrame(klines, columns=["timestamp", "open", "high", "low", "close", "volume", "turnover"])
                df["timestamp"] = pd.to_datetime(df["timestamp"], unit="ms", utc=True)
                df = df.set_index("timestamp")
                for col in ["open", "high", "low", "close", "volume"]:
                    df[col] = pd.to_numeric(df[col], errors="coerce")
                df = df[["open", "high", "low", "close", "volume"]].dropna()
                try:
                    if OHLCV_SOURCE_SUCCESS:
                        OHLCV_SOURCE_SUCCESS.labels(source="toobit").set(1)
                except Exception:
                    pass
                return df
    except Exception as e:
        logger.debug(f"خطا در کندل‌های Toobit برای {symbol}: {e}")
    return pd.DataFrame()

async def fetch_ohlcv_ccxt(session: aiohttp.ClientSession, symbol: str, timeframe: str, limit: int) -> pd.DataFrame:
    class ExchangeContextManager:
        def __init__(self, exchange_name: str):
            self.exchange_name = exchange_name
            self.exchange = None
            
        def __enter__(self):
            exchange_class = getattr(ccxt, self.exchange_name)
            self.exchange = exchange_class({
                'timeout': 15000, 
                'enableRateLimit': True,
                'rateLimit': 1000,  # ✅ اضافه شد
                'sandbox': False,
                'verbose': False    # ✅ اضافه شد
            })
            return self.exchange
            
        def __exit__(self, exc_type, exc_val, exc_tb):
            if self.exchange:
                try:
                    self.exchange.close()
                except Exception:
                    pass

    exchange_names = ["kucoin", "gateio", "bitget", "htx"]
    tf_map = {"15m": "15m", "30m": "30m", "1h": "1h", "4h": "4h", "1d": "1d"}
    ccxt_tf = tf_map.get(timeframe, "1h")

    for exchange_name in exchange_names:
        try:
            with ExchangeContextManager(exchange_name) as exchange:
                ohlcv = exchange.fetch_ohlcv(symbol, ccxt_tf, limit=limit)
                if ohlcv and len(ohlcv) >= 30:
                    df = pd.DataFrame(ohlcv, columns=["timestamp", "open", "high", "low", "close", "volume"])
                    df["timestamp"] = pd.to_datetime(df["timestamp"], unit="ms", utc=True)
                    df = df.set_index("timestamp")
                    for col in ["open", "high", "low", "close", "volume"]:
                        df[col] = pd.to_numeric(df[col], errors="coerce")
                    df = df[["open", "high", "low", "close", "volume"]].dropna()
                    try:
                        if OHLCV_SOURCE_SUCCESS:
                            OHLCV_SOURCE_SUCCESS.labels(source=exchange.id).set(1)
                    except Exception:
                        pass
                    logger.info(f"دریافت OHLCV از {exchange.id} برای {symbol} {timeframe} موفق بود")
                    return df
        except Exception as e:
            logger.debug(f"CCXT {exchange_name} برای {symbol} خطا داد: {e}")
            continue
    return pd.DataFrame()

# --------- DataFetcher (CoinGecko OHLCV fallback) ----------
class DataFetcher:
    def __init__(self, session: aiohttp.ClientSession):
        self.session = session
        self.cache = TTLCache(maxsize=100, ttl=300)

    async def fetch_ohlcv(self, symbol: str, timeframe: str, limit: int) -> pd.DataFrame:
        try:
            coin_map = {
                "BTC/USDT": "bitcoin", "ETH/USDT": "ethereum", "BNB/USDT": "binancecoin",
                "ADA/USDT": "cardano", "SOL/USDT": "solana", "XRP/USDT": "ripple", "DOT/USDT": "polkadot",
            }
            coin_id = coin_map.get(symbol)
            if not coin_id:
                return pd.DataFrame()
            interval_map = {"15m": "15min", "30m": "30min", "1h": "1hour", "4h": "4hour", "1d": "daily"}
            interval = interval_map.get(timeframe, "daily")
            url = f"https://api.coingecko.com/api/v3/coins/{coin_id}/ohlc"
            params = {"vs_currency": "usd", "days": max(1, limit // 24), "interval": interval}
            headers = {"x-cg-demo-api-key": config.COINGECKO_API_KEY} if config.COINGECKO_API_KEY else {}
            data = await http_get(self.session, url, params=params, headers=headers)
            if data and isinstance(data, list) and len(data) > 0:
                df = pd.DataFrame(data, columns=["timestamp", "open", "high", "low", "close"])
                df["timestamp"] = pd.to_datetime(df["timestamp"], unit="ms", utc=True)
                df = df.set_index("timestamp")
                df["volume"] = np.nan
                return df[["open", "high", "low", "close", "volume"]]
        except Exception as e:
            logger.debug(f"DataFetcher - خطا در CoinGecko: {e}")
        return pd.DataFrame()

# --------- Synthetic OHLCV ----------
def generate_synthetic_ohlcv(base_price: float, timeframe: str, periods: int) -> pd.DataFrame:
    try:
        if periods < 2:
            periods = 2
        delta = {
            "15m": timedelta(minutes=15),
            "30m": timedelta(minutes=30),
            "1h": timedelta(hours=1),
            "4h": timedelta(hours=4),
            "1d": timedelta(days=1),
        }.get(timeframe, timedelta(hours=1))
        end = datetime.now(timezone.utc)  # ✅ اصلاح شد
        dates = [end - i * delta for i in range(periods)][::-1]
        prices = [base_price]
        for _ in range(1, periods):
            change = np.random.uniform(-0.02, 0.02)
            prices.append(prices[-1] * (1 + change))
        highs = [p * np.random.uniform(1.00, 1.02) for p in prices]
        lows = [p * np.random.uniform(0.98, 1.00) for p in prices]
        vols = [np.random.uniform(1000, 10000) for _ in prices]
        df = pd.DataFrame(
            {"timestamp": pd.to_datetime(dates, utc=True), "open": prices, "high": highs, "low": lows, "close": prices, "volume": vols}
        )
        df = df.set_index("timestamp")
        logger.info(f"داده مصنوعی OHLCV برای {timeframe} با {periods} دوره تولید شد")
        return df
    except Exception as e:
        logger.error(f"خطا در تولید داده مصنوعی OHLCV: {e}")
        return pd.DataFrame()

# --------- OHLCV orchestration ----------
async def fetch_ohlcv_data(session: aiohttp.ClientSession, symbol: str, timeframe: str, limit: int) -> pd.DataFrame:
    sources = [
        ("MEXC", fetch_mexc_klines),
        ("Toobit", fetch_toobit_klines),
        ("CCXT", fetch_ohlcv_ccxt),
        ("DataFetcher", lambda s, sym, tf, lim: DataFetcher(s).fetch_ohlcv(sym, tf, lim)),
    ]
    logger.info(f"شروع دریافت OHLCV برای {symbol} | تایم‌فریم: {timeframe} | تعداد: {limit}")
    for source_name, fetch_func in sources:
        try:
            df = await fetch_func(session, symbol, timeframe, limit)
            if _validate_ohlcv_data(df, symbol, source_name):
                logger.info(f"منبع {source_name} برای {symbol} موفق بود | تعداد کندل‌ها: {len(df)}")
                return df
            else:
                logger.warning(f"داده‌های {source_name} برای {symbol} نامعتبر است")
        except Exception as e:
            logger.warning(f"خطا در منبع {source_name} برای {symbol}: {str(e)}")
            continue
    logger.error(f"همه منابع برای {symbol} ناموفق بودند")
    _log_failure_metrics(symbol)
    return pd.DataFrame()

def _validate_ohlcv_data(df: pd.DataFrame, symbol: str, source_name: str) -> bool:
    if df.empty:
        logger.debug(f"داده‌های {source_name} برای {symbol} خالی است")
        return False
    if len(df) < 30:
        logger.warning(f"داده‌های {source_name} برای {symbol} ناکافی است: {len(df)} کندل")
        return False
    required = ["open", "high", "low", "close", "volume"]
    if any(c not in df.columns for c in required):
        logger.warning(f"ستون‌های گمشده در {source_name} برای {symbol}")
        return False
    subset = df[required]
    if subset.isna().sum().sum() > 0:
        subset = subset.dropna()
        if len(subset) < 30:
            logger.warning(f"پس از پاک‌سازی NaN، داده‌های {symbol} ناکافی شد: {len(subset)}")
            return False
    if ((df["open"] <= 0).any() or (df["high"] <= 0).any() or (df["low"] <= 0).any() or (df["close"] <= 0).any()):
        logger.warning(f"مقادیر غیرمجاز در OHLC برای {symbol} از {source_name}")
        return False
    if (df["high"] < df["low"]).any():
        logger.warning(f"مقدار high کوچکتر از low در {symbol} از {source_name}")
        return False
    return True

def _log_failure_metrics(symbol: str):
    try:
        if OHLCV_FETCH_FAILURES:
            OHLCV_FETCH_FAILURES.labels(symbol=symbol).inc()
        logger.error(f"ثبت متریک شکست برای {symbol}")
    except Exception as e:
        logger.warning(f"خطا در ثبت متریک‌های شکست: {e}")

# --------- TECHNICAL ANALYSIS ----------
def calculate_stochastic(high: pd.Series, low: pd.Series, close: pd.Series, k_period: int = 14, d_period: int = 3) -> Dict[str, Any]:
    try:
        stoch_k, stoch_d = talib.STOCH(high, low, close, fastk_period=k_period, slowk_period=d_period, slowd_period=d_period)
        current_k = safe_get(stoch_k, -1, 50)
        current_d = safe_get(stoch_d, -1, 50)
        if current_k < 20 and current_d < 20:
            signal, condition = 1.0, "oversold"
        elif current_k > 80 and current_d > 80:
            signal, condition = -1.0, "overbought"
        elif current_k > current_d and current_k > 50:
            signal, condition = 0.5, "bullish"
        elif current_k < current_d and current_k < 50:
            signal, condition = -0.5, "bearish"
        else:
            signal, condition = 0.0, "neutral"
        return {"signal": signal, "k_value": current_k, "d_value": current_d, "condition": condition, "crossover": "bullish" if current_k > current_d else "bearish"}
    except Exception as e:
        logger.debug(f"خطا در محاسبه Stochastic: {e}")
        return {"signal": 0.0, "k_value": 50, "d_value": 50, "condition": "neutral", "crossover": "none"}

def calculate_bollinger_bands(close: pd.Series, period: int = 20, std_dev: int = 2) -> Dict[str, Any]:
    try:
        upper, middle, lower = talib.BBANDS(close, timeperiod=period, nbdevup=std_dev, nbdevdn=std_dev)
        current_price = safe_get(close, -1, 0)
        current_upper = safe_get(upper, -1, current_price)
        current_lower = safe_get(lower, -1, current_price)
        current_middle = safe_get(middle, -1, current_price)
        if current_price <= current_lower:
            signal, position = 1.0, "oversold"
        elif current_price >= current_upper:
            signal, position = -1.0, "overbought"
        else:
            signal, position = 0.0, "inside"
        band_width = (((current_upper - current_lower) / current_middle) * 100) if current_middle else 0
        price_pos = (((current_price - current_lower) / (current_upper - current_lower)) * 100) if (current_upper - current_lower) else 50
        return {"signal": signal, "position": position, "upper_band": current_upper, "middle_band": current_middle, "lower_band": current_lower, "band_width": band_width, "price_position": price_pos}
    except Exception as e:
        logger.warning(f"خطا در محاسبه باندهای بولینگر: {e}")
        cp = safe_get(close, -1, 0)
        return {"signal": 0.0, "position": "inside", "upper_band": cp, "middle_band": cp, "lower_band": cp, "band_width": 0, "price_position": 50}

def calculate_fibonacci_levels(high_price, low_price):
    diff = high_price - low_price
    levels = {
        "0.236": high_price - diff * 0.236,
        "0.382": high_price - diff * 0.382,
        "0.5": high_price - diff * 0.5,
        "0.618": high_price - diff * 0.618,
        "0.786": high_price - diff * 0.786,
        "1.0": low_price,
    }
    tp = next((lvl for lvl in sorted(levels.values()) if lvl > low_price), None)
    return {"levels": levels, "tp_suggestion": tp}

def analyze_volume_profile(volume: pd.Series, lookback_period: int = 20) -> Dict[str, Any]:
    try:
        current_volume = safe_get(volume, -1, 0)
        avg_volume = volume.tail(lookback_period).mean() if len(volume) >= 1 else 0
        volume_ratio = current_volume / avg_volume if avg_volume > 0 else 1.0
        if volume_ratio > 2.0:
            signal, condition = 1.0, "very_high"
        elif volume_ratio > 1.5:
            signal, condition = 0.5, "high"
        elif volume_ratio < 0.5:
            signal, condition = -0.5, "low"
        else:
            signal, condition = 0.0, "normal"
        return {"signal": signal, "current_volume": current_volume, "avg_volume": avg_volume, "volume_ratio": volume_ratio, "condition": condition}
    except Exception as e:
        logger.warning(f"خطا در تحلیل پروفایل حجم: {e}")
        return {"signal": 0.0, "current_volume": 0, "avg_volume": 0, "volume_ratio": 1.0, "condition": "error"}

def analyze_volume_trend(volume: pd.Series) -> Dict[str, Any]:
    try:
        if len(volume) < 3:
            return {"signal": 0.0, "trend": "unknown"}
        v = volume.tail(3).values
        if v[2] > v[1] > v[0]:
            return {"signal": 1.0, "trend": "increasing"}
        if v[2] < v[1] < v[0]:
            return {"signal": -1.0, "trend": "decreasing"}
        return {"signal": 0.0, "trend": "stable"}
    except Exception as e:
        logger.debug(f"خطا در روند حجم: {e}")
        return {"signal": 0.0, "trend": "error"}

def calculate_ichimoku(high: pd.Series, low: pd.Series, close: pd.Series) -> Dict[str, Any]:
    try:
        tenkan_sen = (high.rolling(9).max() + low.rolling(9).min()) / 2
        kijun_sen = (high.rolling(26).max() + low.rolling(26).min()) / 2
        senkou_span_a = ((tenkan_sen + kijun_sen) / 2).shift(26)
        senkou_span_b = ((high.rolling(52).max() + low.rolling(52).min()) / 2).shift(26)
        current_close = safe_get(close, -1, 0)
        cloud_top = max(safe_get(senkou_span_a, -1, current_close), safe_get(senkou_span_b, -1, current_close))
        cloud_bottom = min(safe_get(senkou_span_a, -1, current_close), safe_get(senkou_span_b, -1, current_close))
        if current_close > cloud_top:
            signal, position = 1.0, "above_cloud"
        elif current_close < cloud_bottom:
            signal, position = -1.0, "below_cloud"
        else:
            signal, position = 0.0, "inside_cloud"
        cross_signal = 1.0 if safe_get(tenkan_sen, -1, 0) > safe_get(kijun_sen, -1, 0) else (-1.0 if safe_get(tenkan_sen, -1, 0) < safe_get(kijun_sen, -1, 0) else 0.0)
        cross_type = "bullish" if cross_signal > 0 else ("bearish" if cross_signal < 0 else "neutral")
        return {"signal": signal, "position": position, "tenkan_sen": safe_get(tenkan_sen, -1, 0), "kijun_sen": safe_get(kijun_sen, -1, 0), "senkou_span_a": safe_get(senkou_span_a, -1, 0), "senkou_span_b": safe_get(senkou_span_b, -1, 0), "cross_signal": cross_signal, "cross_type": cross_type}
    except Exception as e:
        logger.debug(f"خطا در ایچیموکو: {e}")
        return {"signal": 0.0, "position": "error", "cross_signal": 0.0, "cross_type": "neutral"}

def detect_advanced_price_spikes(df: pd.DataFrame, threshold: float = 2.0) -> Dict[str, Any]:
    try:
        returns = df["close"].pct_change().dropna()
        mean_return = returns.mean()
        std_return = returns.std()
        current_return = safe_get(returns, -1, 0)
        z_score = (current_return - mean_return) / std_return if std_return > 0 else 0
        if z_score > threshold:
            signal, spike_type = -1.0, "positive_spike"
        elif z_score < -threshold:
            signal, spike_type = 1.0, "negative_spike"
        else:
            signal, spike_type = 0.0, "no_spike"
        return {"signal": signal, "z_score": z_score, "spike_type": spike_type, "current_return": current_return, "volatility": std_return}
    except Exception as e:
        logger.warning(f"خطا در تشخیص اسپایک قیمت: {e}")
        return {"signal": 0.0, "z_score": 0, "spike_type": "no_spike", "current_return": 0, "volatility": 0}

def calculate_candlestick_signal(df: pd.DataFrame) -> Dict[str, Any]:
    try:
        o, h, l, c = df["open"], df["high"], df["low"], df["close"]
        bullish = {
            "hammer": talib.CDLHAMMER(o, h, l, c),
            "inverted_hammer": talib.CDLINVERTEDHAMMER(o, h, l, c),
            "engulfing_bullish": talib.CDLENGULFING(o, h, l, c),
            "morning_star": talib.CDLMORNINGSTAR(o, h, l, c),
            "piercing": talib.CDLPIERCING(o, h, l, c),
            "three_white_soldiers": talib.CDL3WHITESOLDIERS(o, h, l, c),
        }
        bearish = {
            "shooting_star": talib.CDLSHOOTINGSTAR(o, h, l, c),
            "hanging_man": talib.CDLHANGINGMAN(o, h, l, c),
            "engulfing_bearish": talib.CDLENGULFING(o, h, l, c),
            "evening_star": talib.CDLEVENINGSTAR(o, h, l, c),
            "dark_cloud_cover": talib.CDLDARKCLOUDCOVER(o, h, l, c),
            "three_black_crows": talib.CDL3BLACKCROWS(o, h, l, c),
        }
        bullish_count = sum(1 for p in bullish.values() if not p.empty and p.iloc[-1] > 0)
        bearish_count = sum(1 for p in bearish.values() if not p.empty and p.iloc[-1] > 0)
        net = bullish_count - bearish_count
        if net > 1:
            signal, strength = 1.0, "strong_bullish"
        elif net > 0:
            signal, strength = 0.5, "weak_bullish"
        elif net < -1:
            signal, strength = -1.0, "strong_bearish"
        elif net < 0:
            signal, strength = -0.5, "weak_bearish"
        else:
            signal, strength = 0.0, "neutral"
        return {"signal": signal, "bullish_count": bullish_count, "bearish_count": bearish_count, "net_patterns": net, "strength": strength}
    except Exception as e:
        logger.warning(f"خطا در محاسبه الگوهای کندلی: {e}")
        return {"signal": 0.0, "bullish_count": 0, "bearish_count": 0, "net_patterns": 0, "strength": "neutral"}

def generate_technical_summary(analysis: Dict[str, Any]) -> Dict[str, Any]:
    try:
        signals = {
            "macd": analysis.get("macd", {}).get("signal", 0),
            "rsi": analysis.get("rsi", {}).get("signal", 0),
            "stochastic": analysis.get("stochastic", {}).get("signal", 0),
            "bollinger_bands": analysis.get("bollinger_bands", {}).get("signal", 0),
            "ema_cross": analysis.get("ema_cross", {}).get("signal", 0),
            "volume_profile": analysis.get("volume_profile", {}).get("signal", 0),
            "candlestick": analysis.get("candlestick", {}).get("signal", 0),
            "ichimoku": analysis.get("ichimoku", {}).get("signal", 0),
            "parabolic_sar": analysis.get("parabolic_sar", {}).get("signal", 0),
            "williams_r": analysis.get("williams_r", {}).get("signal", 0),
            "obv": analysis.get("obv", {}).get("signal", 0),
        }
        weights = {
            "macd": 0.15, "rsi": 0.15, "stochastic": 0.10, "bollinger_bands": 0.15,
            "ema_cross": 0.15, "volume_profile": 0.10, "candlestick": 0.05,
            "ichimoku": 0.05, "parabolic_sar": 0.05, "williams_r": 0.03, "obv": 0.02,
        }
        total_score = sum(signals[ind] * weights[ind] for ind in weights)
        if total_score > 0.3:
            overall_signal = "BULLISH"
            confidence = min(100, int((total_score + 1) * 50))
        elif total_score < -0.3:
            overall_signal = "BEARISH"
            confidence = min(100, int((abs(total_score) + 1) * 50))
        else:
            overall_signal, confidence = "NEUTRAL", 50
        trend_strength = "strong" if abs(total_score) > 0.6 else ("moderate" if abs(total_score) > 0.3 else "weak")
        return {"overall_signal": overall_signal, "composite_score": round(total_score, 3), "confidence": confidence, "trend_strength": trend_strength}
    except Exception as e:
        logger.warning(f"خطا در تولید خلاصه تکنیکال: {e}")
        return {"overall_signal": "NEUTRAL", "composite_score": 0.0, "confidence": 0, "trend_strength": "weak"}

def enhanced_technical_analysis(df: pd.DataFrame) -> Dict[str, Any]:
    if df is None or df.empty:
        return {"error": "Empty dataframe", "summary": {"overall_signal": "HOLD", "composite_score": 0.0, "confidence": 0}}
    required = ["open", "high", "low", "close"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return {"error": f"Missing columns: {missing}", "summary": {"overall_signal": "HOLD", "composite_score": 0.0, "confidence": 0}}
    analysis: Dict[str, Any] = {}
    score = 0.0
    try:
        macd_line, macd_signal, macd_hist = talib.MACD(df["close"], fastperiod=12, slowperiod=26, signalperiod=9)
        current_hist = safe_get(macd_hist, -1, 0)
        prev_hist = safe_get(macd_hist, -2, 0)
        if current_hist > 0 and current_hist > prev_hist:
            macd_signal_score, macd_trend = 1.0, "bullish_strong"
        elif current_hist < 0 and current_hist < prev_hist:
            macd_signal_score, macd_trend = -1.0, "bearish_strong"
        elif current_hist > 0:
            macd_signal_score, macd_trend = 0.5, "bullish_weak"
        elif current_hist < 0:
            macd_signal_score, macd_trend = -0.5, "bearish_weak"
        else:
            macd_signal_score, macd_trend = 0.0, "neutral"
        analysis["macd"] = {"signal": macd_signal_score, "histogram": float(current_hist), "trend": macd_trend, "line": safe_get(macd_line, -1, 0.0), "signal_line": safe_get(macd_signal, -1, 0.0)}
        score += macd_signal_score * 0.1

        ema_12 = talib.EMA(df["close"], timeperiod=12)
        ema_26 = talib.EMA(df["close"], timeperiod=26)
        ema_cross_signal = 1.0 if safe_get(ema_12, -1, 0) > safe_get(ema_26, -1, 0) else -1.0
        analysis["ema_cross"] = {"signal": ema_cross_signal, "ema_12": safe_get(ema_12, -1, 0.0), "ema_26": safe_get(ema_26, -1, 0.0), "trend": "bullish" if ema_cross_signal > 0 else "bearish"}
        score += ema_cross_signal * 0.1

        sar = talib.SAR(df["high"], df["low"], acceleration=0.02, maximum=0.2)
        sar_signal = 1.0 if safe_get(df["close"], -1, 0) > safe_get(sar, -1, 0) else -1.0
        analysis["parabolic_sar"] = {"signal": sar_signal, "value": safe_get(sar, -1, 0.0), "position": "above" if sar_signal > 0 else "below"}
        score += sar_signal * 0.05

        rsi = talib.RSI(df["close"], timeperiod=14)
        current_rsi = safe_get(rsi, -1, 50.0)
        if current_rsi < 30:
            rsi_signal_score, rsi_condition = 1.0, "oversold"
        elif current_rsi > 70:
            rsi_signal_score, rsi_condition = -1.0, "overbought"
        elif current_rsi < 40:
            rsi_signal_score, rsi_condition = 0.5, "bullish"
        elif current_rsi > 60:
            rsi_signal_score, rsi_condition = -0.5, "bearish"
        else:
            rsi_signal_score, rsi_condition = 0.0, "neutral"
        analysis["rsi"] = {"signal": rsi_signal_score, "value": float(current_rsi), "condition": rsi_condition}
        score += rsi_signal_score * 0.1

        analysis["stochastic"] = calculate_stochastic(df["high"], df["low"], df["close"])

        wr = talib.WILLR(df["high"], df["low"], df["close"], timeperiod=14)
        current_wr = safe_get(wr, -1, -50)
        if current_wr > -20:
            wr_signal, wr_condition = -1.0, "overbought"
        elif current_wr < -80:
            wr_signal, wr_condition = 1.0, "oversold"
        else:
            wr_signal, wr_condition = 0.0, "neutral"
        analysis["williams_r"] = {"signal": wr_signal, "value": float(current_wr), "condition": wr_condition}
        score += wr_signal * 0.05

        analysis["bollinger_bands"] = calculate_bollinger_bands(df["close"])

        atr_series = talib.ATR(df["high"], df["low"], df["close"], timeperiod=14)
        atr_value = safe_get(atr_series, -1, 0.0)
        price = safe_get(df["close"], -1, 1.0)
        atr_percentage = (atr_value / price) * 100 if price > 0 else 0.0
        if atr_percentage > 3:
            atr_signal, atr_condition = -1.0, "high_volatility"
        elif atr_percentage < 1:
            atr_signal, atr_condition = 0.5, "low_volatility"
        else:
            atr_signal, atr_condition = 0.0, "normal_volatility"
        analysis["atr"] = {"signal": atr_signal, "value": float(atr_value), "percentage": float(atr_percentage), "condition": atr_condition}
        score += atr_signal * 0.05

        vol = df["volume"].fillna(df["volume"].median() if not df["volume"].empty else 1)
        analysis["volume_profile"] = analyze_volume_profile(vol)
        analysis["volume_trend"] = analyze_volume_trend(vol)

        obv = talib.OBV(df["close"], vol.fillna(0))
        if not obv.empty and len(obv) > 1:
            obv_trend = safe_get(obv, -1, 0) - safe_get(obv, -2, 0)
            obv_signal = 1.0 if obv_trend > 0 else (-1.0 if obv_trend < 0 else 0.0)
        else:
            obv_signal = 0.0
        analysis["obv"] = {"signal": obv_signal, "trend": ("rising" if obv_signal > 0 else ("falling" if obv_signal < 0 else "flat"))}
        score += obv_signal * 0.05

        analysis["ichimoku"] = calculate_ichimoku(df["high"], df["low"], df["close"])
        analysis["price_spike"] = detect_advanced_price_spikes(df)
        analysis["candlestick"] = calculate_candlestick_signal(df)
        analysis["summary"] = generate_technical_summary(analysis)

        swing_high = float(df["high"].max())
        swing_low = float(df["low"].min())
        fibo = calculate_fibonacci_levels(swing_high, swing_low)
        analysis.setdefault("indicators", {})
        analysis["indicators"]["fibonacci"] = fibo["levels"]
        analysis["indicators"]["tp_fibonacci"] = fibo["tp_suggestion"]

        current_price = safe_get(df["close"], -1, 0)
        for level_value in fibo["levels"].values():
            if current_price and abs(current_price - level_value) / current_price < 0.01:
                score += 0.08
                break

        analysis["composite_score"] = float(score)
        return analysis
    except Exception as e:
        logger.warning(f"خطا در تحلیل تکنیکال پیشرفته: {e}")
        return {"summary": {"overall_signal": "NEUTRAL", "composite_score": 0.0, "confidence": 0}}

# --------- RULE ENGINE (UNIFIED STRATEGY) ----------
class UnifiedStrategy:
    def __init__(self, symbol: str, timeframe: str, side: str, entry: float, sl: float, tp: float, confidence: float, rule: str):
        self.symbol = symbol
        self.timeframe = timeframe
        self.side = side  # "Buy" یا "Sell"
        self.entry = entry
        self.sl = sl
        self.tp = tp
        self.confidence = confidence  # 0..100
        self.rule = rule  # نام قانون/الگو

class SimpleRuleSignal:
    def __init__(self, side: str, entry: float, sl: Optional[float], tp: Optional[float], confidence: float, rule: str):
        self.side = side
        self.entry = entry
        self.sl = sl
        self.tp = tp
        self.confidence = confidence
        self.rule = rule

def generate_rule_signals(df_rule: pd.DataFrame, symbol: str, timeframe: str = "") -> List[UnifiedStrategy]:
    """
    تولید سیگنال‌های ساده مبتنی بر کراس EMA و RSI با SL/TP پایه.
    df_rule دارای ستون‌های: Date, Open, High, Low, Close, Volume
    """
    signals: List[UnifiedStrategy] = []
    try:
        close = df_rule["Close"]
        # EMAها
        ema_short = talib.EMA(close, timeperiod=12)
        ema_long = talib.EMA(close, timeperiod=26)
        # RSI
        rsi = talib.RSI(close, timeperiod=14)

        # کراس صعودی/نزولی
        bullish_cross = ema_short.iloc[-1] > ema_long.iloc[-1] and ema_short.iloc[-2] <= ema_long.iloc[-2]
        bearish_cross = ema_short.iloc[-1] < ema_long.iloc[-1] and ema_short.iloc[-2] >= ema_long.iloc[-2]

        # شرط RSI برای فیلتر نویز
        rsi_bull_ok = rsi.iloc[-1] < 65  # از ناحیه اشباع خرید دور باشد
        rsi_bear_ok = rsi.iloc[-1] > 35  # از ناحیه اشباع فروش دور باشد

        entry = float(close.iloc[-1])
        # SL/TP پایه بر اساس نوسان اخیر
        swing_high = float(df_rule["High"].rolling(20).max().iloc[-1])
        swing_low = float(df_rule["Low"].rolling(20).min().iloc[-1])
        atr_proxy = float((df_rule["High"] - df_rule["Low"]).tail(14).mean())

        if bullish_cross and rsi_bull_ok:
            sl = max(swing_low - 0.5 * atr_proxy, entry * 0.98)
            tp = entry + (entry - sl) * 1.5
            conf = 70.0
            signals.append(UnifiedStrategy(symbol, timeframe or "", "Buy", entry, sl, tp, conf, "EMA12/26 + RSI"))

        elif bearish_cross and rsi_bear_ok:
            sl = min(swing_high + 0.5 * atr_proxy, entry * 1.02)
            tp = entry - (sl - entry) * 1.5
            conf = 70.0
            signals.append(UnifiedStrategy(symbol, timeframe or "", "Sell", entry, sl, tp, conf, "EMA12/26 + RSI"))

        # در صورت نبود کراس، اگر RSI خیلی افراطی بود، سیگنال ضعیف ایجاد کن
        else:
            if rsi.iloc[-1] < 30:
                sl = max(swing_low - 0.5 * atr_proxy, entry * 0.98)
                tp = entry + (entry - sl) * 1.2
                signals.append(UnifiedStrategy(symbol, timeframe or "", "Buy", entry, sl, tp, 55.0, "RSI<30"))
            elif rsi.iloc[-1] > 70:
                sl = min(swing_high + 0.5 * atr_proxy, entry * 1.02)
                tp = entry - (sl - entry) * 1.2
                signals.append(UnifiedStrategy(symbol, timeframe or "", "Sell", entry, sl, tp, 55.0, "RSI>70"))

    except Exception as e:
        logger.debug(f"generate_rule_signals error for {symbol} {timeframe}: {e}")

    return signals

def to_rule_df(df_coingecko: pd.DataFrame) -> pd.DataFrame:
    """تبدیل DataFrame از فرمت CoinGecko به فرمت rule-compatible"""
    df = df_coingecko.copy()
    df = df.rename(columns={
        'open': 'Open', 
        'high': 'High', 
        'low': 'Low', 
        'close': 'Close', 
        'volume': 'Volume'
    })
    df['Date'] = df.index
    df = df.reset_index(drop=True)
    df = df.sort_values('Date').reset_index(drop=True)
    
    # اطمینان از وجود تمام ستون‌های لازم
    for col in ['Open', 'High', 'Low']:
        if col not in df.columns:
            df[col] = df['Close']
    if 'Volume' not in df.columns:
        df['Volume'] = np.nan
        
    return df[['Date', 'Open', 'High', 'Low', 'Close', 'Volume']]

# --------- AdvancedTechnicalAnalyzer با پشتیبانی از Rule Engine ----------
class AdvancedTechnicalAnalyzer:
    def __init__(self, df: pd.DataFrame = None):
        self.df = df
        self.base_weights = {
            'macd': 0.15, 'rsi': 0.10, 'fibonacci': 0.10, 'volume': 0.10,
            'atr': 0.10, 'candlestick': 0.10, 'ichimoku': 0.10, 'divergence': 0.10,
            'adx': 0.05, 'bollinger': 0.05, 'ema_cross': 0.05, 'news': 0.10,
            'harmonic': 0.05, 'obv': 0.04, 'vwap': 0.04, 'supertrend': 0.05, 'psar': 0.04
        }
        self.cache = TTLCache(maxsize=200, ttl=300)

    def set_data(self, df: pd.DataFrame):
        self.df = df

    def cleanup(self):
        self.df = None
        gc.collect()

    def adjust_weights_dynamically(self, price_reliability: float, news_reliability: float) -> Dict[str, float]:
        weights = self.base_weights.copy()
        weights['news'] *= (1 + news_reliability)
        if price_reliability > 0.7:
            weights['macd'] += 0.03
            weights['rsi'] += 0.03
            weights['ema_cross'] += 0.02
        total = sum(weights.values())
        return {k: v / total for k, v in weights.items()}

    def comprehensive_analysis(self, price_reliability: float = 1.0, news_reliability: float = 1.0, news_score: float = 0.0) -> Dict[str, Any]:
        if self.df is None or self.df.empty:
            return {"signal": "HOLD", "score": 0.0, "confidence": 0, "indicators": {}, "summary": {"overall_signal": "HOLD", "confidence": 0}}
        
        # استفاده از تحلیل تکنیکال موجود
        analysis = enhanced_technical_analysis(self.df)
        composite_score = analysis["summary"].get("composite_score", analysis.get("composite_score", 0.0))
        
        reliability_factor = (price_reliability + news_reliability) / 2
        final_score = composite_score * reliability_factor
        
        if final_score > 0.3:
            signal, confidence = "STRONG_BUY", min(100, int((final_score + 1) * 50))
        elif final_score > 0.1:
            signal, confidence = "BUY", min(80, int((final_score + 1) * 40))
        elif final_score < -0.3:
            signal, confidence = "STRONG_SELL", min(100, int((abs(final_score) + 1) * 50))
        elif final_score < -0.1:
            signal, confidence = "SELL", min(80, int((abs(final_score) + 1) * 40))
        else:
            signal, confidence = "HOLD", 50
            
        return {
            "signal": signal,
            "score": round(final_score, 3),
            "confidence": confidence,
            "indicators": analysis,
            "summary": analysis["summary"],
            "reliability": {"price": round(price_reliability, 3), "news": round(news_reliability, 3), "overall": round(reliability_factor, 3)},
        }

    def calculate_sl_tp(self, signal: str, current_price: float, atr: float = None) -> Tuple[float, float]:
        if atr is None:
            try:
                atr = talib.ATR(self.df["high"], self.df["low"], self.df["close"], timeperiod=14).iloc[-1]
            except Exception:
                atr = current_price * 0.02
        if signal in ["BUY", "STRONG_BUY"]:
            sl = current_price - (atr * 1.5)
            tp = current_price + (atr * 2.5)
        elif signal in ["SELL", "STRONG_SELL"]:
            sl = current_price + (atr * 1.5)
            tp = current_price - (atr * 2.5)
        else:
            sl = current_price * 0.99
            tp = current_price * 1.01
        return float(sl), float(tp)

# --------- ADAPTIVE WEIGHTS FUNCTION ----------
def adapt_weights_from_history(analyzer: AdvancedTechnicalAnalyzer, lookback=500):
    try:
        conn = sqlite3.connect(config.DB_PATH)
        df = pd.read_sql_query("SELECT * FROM signals ORDER BY id DESC LIMIT ?", conn, params=(lookback,))
        conn.close()
        if df.empty:
            return
        avg_conf_buy = df[df['signal'].str.contains('BUY', na=False)]['confidence'].mean()
        avg_conf_sell = df[df['signal'].str.contains('SELL', na=False)]['confidence'].mean()
        if pd.notnull(avg_conf_buy) and avg_conf_buy > 60:
            analyzer.base_weights['supertrend'] = analyzer.base_weights.get('supertrend', 0.05) + 0.01
            analyzer.base_weights['ema_cross'] = analyzer.base_weights.get('ema_cross', 0.05) + 0.01
        if pd.notnull(avg_conf_sell) and avg_conf_sell > 60:
            analyzer.base_weights['psar'] = analyzer.base_weights.get('psar', 0.04) + 0.01
        total = sum(analyzer.base_weights.values())
        analyzer.base_weights = {k: v / total for k, v in analyzer.base_weights.items()}
    except Exception as e:
        logger.debug(f"Adaptive weight update failed: {e}")

# --------- ML SIGNAL ENHANCER ----------
class MLSignalEnhancer:
    def __init__(self):
        self.is_trained = config.USE_ML_PREDICTIONS

    def predict_signal(self, technical_data: Dict, price_data: pd.DataFrame) -> Dict:
        if not self.is_trained:
            return {"ml_signal": "HOLD", "ml_confidence": 0.0, "anomaly_score": 0.0, "ensemble_confidence": 0.0}
        try:
            rsi_val = technical_data.get("rsi", {}).get("value", 50)
            macd_hist = technical_data.get("macd", {}).get("histogram", 0)
            vol_ratio = technical_data.get("volume_profile", {}).get("volume_ratio", 1.0)
            ml_score = 0
            if rsi_val < 30:
                ml_score += 2
            elif rsi_val > 70:
                ml_score -= 2
            ml_score += 1 if macd_hist > 0 else -1
            if vol_ratio > 1.5:
                ml_score += 1
            if ml_score >= 2:
                ml_signal, ml_conf = "BUY", min(0.95, 0.8 + (ml_score - 2) * 0.1)
            elif ml_score <= -2:
                ml_signal, ml_conf = "SELL", min(0.95, 0.8 + abs(ml_score + 2) * 0.1)
            else:
                ml_signal, ml_conf = "HOLD", 0.5
            if not price_data.empty:
                price_std = price_data["close"].std()
                current_price = price_data["close"].iloc[-1]
                avg_price = price_data["close"].mean()
                anomaly_score = (min(abs(current_price - avg_price) / price_std / 3, 1.0) if price_std and price_std > 0 else 0.1)
            else:
                anomaly_score = 0.1
            return {"ml_signal": ml_signal, "ml_confidence": float(ml_conf), "anomaly_score": float(anomaly_score), "ensemble_confidence": float(ml_conf * 0.8 + (1 - anomaly_score) * 0.2)}
        except Exception as e:
            logger.warning(f"خطا در پیش‌بینی ML: {e}")
            return {"ml_signal": "HOLD", "ml_confidence": 0.0, "anomaly_score": 0.0, "ensemble_confidence": 0.0}

# --------- ENHANCED ANALYZE SYMBOL WITH RULE INTEGRATION ----------
async def analyze_symbol(
    symbol: str,
    timeframe: str,
    session: aiohttp.ClientSession,
    news_cache: TTLCache,
    price_cache: Dict[str, Tuple[Optional[float], float]]
) -> Optional[Dict]:
    """نسخه بهبود یافته analyze_symbol با rule integration"""
    logger.info(f"Starting analysis with rules {symbol} on {timeframe}")
    
    try:
        # دریافت قیمت
        if symbol in price_cache:
            current_price, price_reliability = price_cache[symbol]
        else:
            current_price, price_reliability = await fetch_price_best(session, symbol)

        # دریافت اخبار
        cache_key = f"news_{symbol}"
        if cache_key in news_cache:
            news_count, news_reliability, news_score = news_cache[cache_key]
        else:
            news_count, news_reliability, news_score = await fetch_news_total(session, symbol)
            news_cache[cache_key] = (news_count, news_reliability, news_score)

        # دریافت داده OHLCV
        df_ohlcv = await fetch_ohlcv_data(session, symbol, timeframe, 150)
        
        # Fallback به DataFetcher اگر داده ناکافی است
        if df_ohlcv.empty or len(df_ohlcv) < 50:
            logger.warning(f"داده OHLCV ناکافی برای {symbol}، استفاده از DataFetcher...")
            df_ohlcv = await DataFetcher(session).fetch_ohlcv(symbol, timeframe, 150)

        if df_ohlcv.empty or current_price is None:
            logger.warning(f"داده ناکافی برای {symbol} {timeframe}")
            return None

        # تحلیل تکنیکال پیشرفته
        analyzer = AdvancedTechnicalAnalyzer(df_ohlcv)
        adapt_weights_from_history(analyzer)
        analysis = analyzer.comprehensive_analysis(price_reliability, news_reliability, news_score)

        # محاسبه SL/TP تکنیکال
        tech_sl, tech_tp = analyzer.calculate_sl_tp(analysis['signal'], current_price)
        
        # Fallback-safe برای SL/TP اگر None باشد
        entry_fallback = float(df_ohlcv['close'].iloc[-1]) if not df_ohlcv.empty else current_price
        if tech_sl is None:
            tech_sl = entry_fallback * 0.98 if analysis['signal'] in ['BUY', 'STRONG_BUY'] else entry_fallback * 1.02
        if tech_tp is None:
            tech_tp = entry_fallback * 1.02 if analysis['signal'] in ['BUY', 'STRONG_BUY'] else entry_fallback * 0.98

        # Rule-based signals
        df_rule = to_rule_df(df_ohlcv)
        rule_signals = generate_rule_signals(df_rule, symbol=symbol, timeframe=timeframe)

        # تلفیق نتایج
        entry_price = current_price
        rule_side = rule_entry = rule_sl = rule_tp = rule_conf = rule_rule = None

        if rule_signals:
            rs = rule_signals[0]
            rule_side = getattr(rs, 'side', None)
            rule_entry = getattr(rs, 'entry', None)
            if isinstance(rule_entry, (int, float)) and rule_entry > 0:
                entry_price = float(rule_entry)
            
            # SL/TP از Rule با fallback ایمن
            rsl = getattr(rs, 'sl', None)
            rtp = getattr(rs, 'tp', None)
            if rsl is None:
                rsl = entry_price * (0.98 if rule_side == "Buy" else 1.02)
            if rtp is None:
                rtp = entry_price * (1.05 if rule_side == "Buy" else 0.95)
            rule_sl, rule_tp = float(rsl), float(rtp)
            rule_conf = float(getattr(rs, 'confidence', 0.0))
            rule_rule = getattr(rs, 'rule', None)

        # نتیجه نهایی با اولویت rule-based signals
        result = {
            'symbol': symbol,
            'timeframe': timeframe,
            'signal': analysis['signal'],
            'score': analysis['score'],
            'confidence': analysis['confidence'],
            'indicators': analysis['indicators'],
            'harmonic_patterns': analysis.get('harmonic_patterns', []),
            'current_price': current_price,
            'entry_price': entry_price,
            # اولویت با Rule اگر موجود باشد
            'stop_loss': rule_sl if rule_sl is not None else tech_sl,
            'take_profit': rule_tp if rule_tp is not None else tech_tp,
            'news_count': news_count,
            'price_reliability': price_reliability,
            'news_reliability': news_reliability,
            'news_score': news_score,
            'timestamp': datetime.now().isoformat(),
            'rule_side': rule_side,
            'rule_entry': rule_entry,
            'rule_sl': rule_sl,
            'rule_tp': rule_tp,
            'rule_confidence': rule_conf,
            'rule_name': rule_rule,
            'ts': f"{symbol}-{timeframe}-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            'price': entry_price,
            'sl': rule_sl if rule_sl is not None else tech_sl,
            'tp': rule_tp if rule_tp is not None else tech_tp,
            'price_rel': price_reliability,
            'news_rel': news_reliability
        }

        # لاگ کردن نتایج
        if result['signal'] != 'HOLD' or rule_side:
            logger.info(
                f"Strong signal: {symbol} {timeframe} → {result['signal']} | "
                f"Conf={result['confidence']:.1f}% | "
                f"Price={current_price:.6f} | Rule: {rule_rule or '—'}"
            )
        else:
            logger.debug(f"{symbol} {timeframe} → HOLD")

        return result

    except Exception as e:
        logger.error(f"Error analyzing {symbol} {timeframe} with rules: {e}", exc_info=True)
        return None

# --------- TELEGRAM ----------
async def send_telegram_message(session: aiohttp.ClientSession, text: str, max_retries: int = 3) -> bool:
    if not config.TELEGRAM_ENABLED:
        return False
    for attempt in range(max_retries):
        try:
            url = f"https://api.telegram.org/bot{config.TELEGRAM_TOKEN}/sendMessage"
            payload = {"chat_id": config.TELEGRAM_CHAT_ID, "text": text, "parse_mode": "Markdown", "disable_web_page_preview": True}
            timeout = ClientTimeout(total=10 + attempt * 5)
            async with session.post(url, json=payload, timeout=timeout) as resp:
                if resp.status == 200:
                    logger.info("پیام تلگرام با موفقیت ارسال شد")
                    return True
                else:
                    error_text = await resp.text()
                    logger.warning(f"پاسخ تلگرام: {resp.status} - {error_text}")
        except asyncio.TimeoutError:
            logger.warning(f"تایم‌اوت تلگرام (تلاش {attempt + 1}/{max_retries})")
        except aiohttp.ClientError as e:
            logger.warning(f"خطای شبکه تلگرام (تلاش {attempt + 1}/{max_retries}): {e}")
        except Exception as e:
            logger.warning(f"خطای غیرمنتظره تلگرام (تلاش {attempt + 1}/{max_retries}): {e}")
        if attempt < max_retries - 1:
            await asyncio.sleep(2**attempt)
    return False

def format_top_signals(signals: List[Dict[str, Any]], max_signals: int = 5) -> str:
    if not signals:
        return "📊 *هیچ سیگنالی موجود نیست*"
    
    # اولویت‌بندی سیگنال‌ها: rule-based اولویت دارند
    def signal_priority(signal):
        priority = 0
        # سیگنال‌های rule-based اولویت بالاتر
        if signal.get('rule_side'):
            priority += 100
        # سیگنال‌های STRONG اولویت بالاتر
        if "STRONG" in str(signal.get("signal", "")):
            priority += 50
        # اعتماد و امتیاز
        priority += safe_float(signal.get("confidence", 0)) + safe_float(signal.get("score", 0)) * 10
        return priority
    
    sorted_signals = sorted(signals, key=signal_priority, reverse=True)[:max_signals]
    
    lines = ["🚀 *سیگنال‌های برتر با قوانین* 🚀"]
    for s in sorted_signals:
        entry = s.get("entry_price", None)
        sl = s.get("stop_loss", None)
        tp = s.get("take_profit", None)
        
        # اطلاعات rule-based
        rule_info = ""
        if s.get('rule_side'):
            rule_info = f"📋 *Rule: {s.get('rule_side', '-')}* @ {fmt_num(s.get('rule_entry', ''))}\n"
            rule_info += f"📝 {s.get('rule_name', '-')} (Conf: {safe_float(s.get('rule_confidence', 0)):.1f}%)\n"
        
        lines.append(
            f"*{s.get('symbol','')} {s.get('timeframe','')}* → `{s.get('signal','')}`\n"
            f"{rule_info}"
            f"🎯 امتیاز={safe_float(s.get('score',0)):.3f} | اعتماد={safe_float(s.get('confidence',0)):.1f}%\n"
            f"💰 ورود={fmt_num(entry)} | SL={fmt_num(sl)} | TP={fmt_num(tp)}\n"
            f"📊 اندازه موقعیت={fmt_num(s.get('position_size', 0))}\n"
            f"📰 اخبار={int(s.get('news_count',0))} | "
            f"🔢 قابلیت اتکای قیمت={safe_float(s.get('price_reliability',0))*100:.1f}% | "
            f"📈 قابلیت اتکای خبر={safe_float(s.get('news_reliability',0))*100:.1f}%\n"
            f"🎭 الگوهای هارمونیک={len(s.get('harmonic_patterns', []))}"
        )
    
    # آمار rule-based
    rule_based_count = len([x for x in signals if x.get('rule_side')])
    total_signals = len(signals)
    strong_signals = len([x for x in signals if "STRONG" in str(x.get("signal", ""))])
    
    lines.append(f"\n📈 خلاصه: {total_signals} سیگنال | {strong_signals} سیگنال قوی | {rule_based_count} سیگنال مبتنی بر قانون")
    lines.append(f"🕒 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    return "\n\n".join(lines)

async def send_telegram_with_fallback(session: aiohttp.ClientSession, message: str, max_retries: int = 3) -> bool:
    ok = await send_telegram_message(session, message, max_retries)
    if ok:
        return True
    try:
        os.makedirs("logs", exist_ok=True)
        filename = "logs/pending_telegram_messages.txt"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(filename, "a", encoding="utf-8") as f:
            f.write("\n" + "=" * 50 + "\n")
            f.write(f"📨 پیام ذخیره شده در: {timestamp}\n")
            f.write("=" * 50 + "\n")
            f.write(f"{message}\n")
            f.write("=" * 50 + "\n")
        logger.info(f"پیام در فایل ذخیره شد: {filename}")
        return True
    except Exception as e:
        logger.error(f"خطا در ذخیره پیام در فایل: {e}")
        return False

def format_analysis_summary(total_signals: int, analysis_time: float, top_signals: List[Dict]) -> str:
    lines = [
        "📊 خلاصه تحلیل بازار با قوانین",
        f"⏱ زمان تحلیل: `{analysis_time:.2f}` ثانیه",
        f"📈 تعداد سیگنال‌ها: `{total_signals}`",
    ]
    
    # آمار rule-based
    rule_based_count = len([x for x in top_signals if x.get('rule_side')])
    if rule_based_count > 0:
        lines.append(f"📋 سیگنال‌های مبتنی بر قانون: `{rule_based_count}`")
    
    lines.append(f"🕒 تاریخ: `{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`")
    
    if top_signals:
        lines.append("\n🎯 سیگنال‌های برتر:")
        for i, signal in enumerate(top_signals[:3], 1):
            rule_marker = " 📋" if signal.get('rule_side') else ""
            lines.append(f"{i}. `{signal.get('symbol','')} {signal.get('timeframe','')}` → {signal.get('signal','')}{rule_marker} ({safe_float(signal.get('confidence',0)):.1f}%)")
    
    return "\n".join(lines)

async def send_pending_messages(session: aiohttp.ClientSession):
    try:
        filename = "logs/pending_telegram_messages.txt"
        if not os.path.exists(filename):
            return
        with open(filename, "r", encoding="utf-8") as f:
            content = f.read().strip()
        if not content:
            return
        messages = content.split("=" * 50)
        sent_any = False
        for message in messages:
            message = message.strip()
            if message and "پیام ذخیره شده در:" in message:
                body_lines = [line for line in message.split("\n") if not any(x in line for x in ["پیام ذخیره شده در:", "=", "📨"])]
                main_message = "\n".join(body_lines).strip()
                if main_message:
                    success = await send_telegram_message(session, main_message)
                    if success:
                        sent_any = True
                        logger.info("پیام ذخیره‌شده با موفقیت ارسال شد")
        if sent_any:
            os.remove(filename)
            logger.info("فایل پیام‌های ذخیره‌شده پاک شد")
    except Exception as e:
        logger.warning(f"خطا در ارسال پیام‌های ذخیره‌شده: {e}")

# --------- HEALTH CHECK & METRICS SERVER ----------
class HealthMonitor:
    def __init__(self):
        self.start_time = datetime.now(timezone.utc)  # ✅ اصلاح شد
        self.cache = PRICE_CACHE

    def get_health_status(self):
        return {
            "status": "healthy",
            "timestamp": datetime.now(timezone.utc).isoformat(),  # ✅ اصلاح شد
            "uptime_seconds": (datetime.now(timezone.utc) - self.start_time).total_seconds(),
            "cache_hit_rate": round(self.cache.get_hit_rate() * 100, 2),
            "cache_size": len(self.cache.cache),
            "memory_usage_mb": self.get_memory_usage(),
        }

    def get_memory_usage(self):
        try:
            process = psutil.Process()
            return round(process.memory_info().rss / 1024 / 1024, 2)
        except Exception:
            return 0.0

health_monitor = HealthMonitor()

async def health_check(request):
    health_data = health_monitor.get_health_status()
    return web.json_response(health_data)

async def metrics_handler(request):
    try:
        metrics_data = generate_latest()
        return web.Response(body=metrics_data, content_type="text/plain")
    except Exception as e:
        return web.json_response({"error": str(e)}, status=500)

async def start_health_server():
    app = web.Application()
    app.router.add_get("/health", health_check)
    app.router.add_get("/metrics", metrics_handler)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", 8080)
    await site.start()
    logger.info("سرور سلامت و متریک‌ها روی http://0.0.0.0:8080 شروع شد")
    return runner

# --------- EXCEL REPORTER ----------
class ExcelReporter:
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_report(self, signals: List[Dict[str, Any]]):
        if not signals:
            logger.info("سیگنالی برای خروجی اکسل موجود نیست")
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.output_dir, f"crypto_signals_{timestamp}.xlsx")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "سیگنال‌ها"
            headers = [
                "نماد",
                "تایم‌فریم",
                "سیگنال",
                "اعتماد (%)",
                "امتیاز",
                "قیمت ورود",
                "حد ضرر",
                "حد سود",
                "تعداد خبر",
                "قابلیت اتکای قیمت (%)",
                "قابلیت اتکای خبر (%)",
                "امتیاز خبر",
                "تاریخ تحلیل",
                "قانون",
                "ورود قانون",
                "SL قانون", 
                "TP قانون",
                "اعتماد قانون (%)"
            ]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            buy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            sell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row, s in enumerate(signals, start=2):
                sig = s.get("signal", "")
                ws.cell(row=row, column=1, value=s.get("symbol", ""))
                ws.cell(row=row, column=2, value=s.get("timeframe", ""))
                c3 = ws.cell(row=row, column=3, value=sig)
                ws.cell(row=row, column=4, value=round(safe_float(s.get("confidence", 0)), 2))
                ws.cell(row=row, column=5, value=round(safe_float(s.get("score", 0)), 6))
                ws.cell(row=row, column=6, value=safe_float(s.get("entry_price", s.get("current_price", 0))))
                ws.cell(row=row, column=7, value=safe_float(s.get("stop_loss", s.get("sl", 0))))
                ws.cell(row=row, column=8, value=safe_float(s.get("take_profit", s.get("tp", 0))))
                ws.cell(row=row, column=9, value=int(s.get("news_count", 0)))
                pr = s.get("price_reliability", s.get("price_rel", 0))
                nr = s.get("news_reliability", s.get("news_rel", 0))
                ws.cell(row=row, column=10, value=round(safe_float(pr) * 100, 2))
                ws.cell(row=row, column=11, value=round(safe_float(nr) * 100, 2))
                ws.cell(row=row, column=12, value=safe_float(s.get("news_score", 0)))
                ws.cell(row=row, column=13, value=s.get("timestamp", ""))
                # اطلاعات rule-based
                ws.cell(row=row, column=14, value=s.get("rule_name", ""))
                ws.cell(row=row, column=15, value=safe_float(s.get("rule_entry", 0)))
                ws.cell(row=row, column=16, value=safe_float(s.get("rule_sl", 0)))
                ws.cell(row=row, column=17, value=safe_float(s.get("rule_tp", 0)))
                ws.cell(row=row, column=18, value=round(safe_float(s.get("rule_confidence", 0)), 2))
                
                if "BUY" in str(sig):
                    c3.fill = buy_fill
                elif "SELL" in str(sig):
                    c3.fill = sell_fill

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            wb.save(filename)
            logger.info(f"فایل اکسل ذخیره شد: {filename}")
        except Exception as e:
            logger.error(f"خطا در ساخت اکسل: {e}")

# --------- MAIN ANALYSIS ----------
async def main_analysis(send_top_to_telegram: bool = True) -> List[Dict[str, Any]]:
    logger.info("شروع تحلیل جامع بازار رمزارز...")
    init_db()
    update_database_schema()
    news_cache = TTLCache(maxsize=50, ttl=600)

    async with aiohttp.ClientSession(connector=TCPConnector(limit=20)) as session:
        # ارسال پیام‌های در صف (در صورت وجود)
        await send_pending_messages(session)

        price_cache: Dict[str, Tuple[Optional[float], float]] = {}
        for sym in config.SYMBOLS:
            try:
                result = await fetch_price_weighted(session, sym)
                if result and len(result) >= 2:
                    price, rel, _ = result
                    price_cache[sym] = (price, rel)
                    if price is not None:
                        logger.info(f"قیمت {sym}: {fmt_num(price)} (اتکا: {rel*100:.1f}%)")
                    else:
                        logger.warning(f"قیمت {sym} ناموفق بود")
                else:
                    price_cache[sym] = (None, 0.0)
                    logger.warning(f"خطا در دریافت قیمت {sym}: نتیجه نامعتبر")
            except Exception as e:
                logger.error(f"خطا در دریافت قیمت {sym}: {e}")
                price_cache[sym] = (None, 0.0)

        logger.info("💰 دریافت قیمت‌ها کامل شد. شروع تحلیل نمادها...")

        semaphore = asyncio.Semaphore(10)

        async def analyze_with_limit(symbol, timeframe):
            async with semaphore:
                logger.info(f"🎯 شروع تحلیل {symbol} {timeframe}")
                result = await analyze_symbol(symbol, timeframe, session, news_cache, price_cache)
                logger.info(f"✅ پایان تحلیل {symbol} {timeframe} - نتیجه: {result.get('signal') if result else 'None'}")
                return result

        tasks = [analyze_with_limit(sym, tf) for sym in config.SYMBOLS for tf in config.TIMEFRAMES]
        logger.info(f"📊 تعداد تسک‌های تحلیل: {len(tasks)}")
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        logger.info(f"📋 جمع‌آوری نتایج کامل شد: {len(results)} نتیجه")

        # 🔥 اعمال فیلترهای پیشرفته
        raw_signals = [
            r for r in results
            if isinstance(r, dict) and r.get("signal") and (r.get("signal") != "HOLD" or r.get('rule_side'))
        ]
        
        # اعمال فیلترهای پیشرفته
        filtered_signals = apply_final_filters(raw_signals, config.INITIAL_BALANCE)
        
        # گزارش تحلیل
        log_signal_analysis(raw_signals, filtered_signals)
        
        logger.info(f"🎯 سیگنال‌های معتبر یافت شده: {len(filtered_signals)}")

        # برای دیباگ، همه نتایج را لاگ کنید
        for i, result in enumerate(results):
            if isinstance(result, dict):
                logger.info(f"🔍 نتیجه {i}: {result.get('symbol')} {result.get('timeframe')} -> {result.get('signal')} (اطمینان: {result.get('confidence')}%)")
            elif isinstance(result, Exception):
                logger.error(f"❌ خطا در تسک {i}: {result}")
        
        filtered_signals.sort(key=lambda x: safe_float(x.get("confidence", 0)), reverse=True)
        top_signals = filtered_signals[:3]

        for s in filtered_signals:
            try:
                log_signal_to_db(s)
                logger.info(f"💾 سیگنال ذخیره شد: {s.get('symbol')} {s.get('timeframe')}")
            except Exception as e:
                logger.error(f"خطا در ذخیره سیگنال {s.get('symbol')}: {e}")

        try:
            reporter = ExcelReporter(config.OUTPUT_DIR)
            reporter.generate_report(filtered_signals)
            logger.info("گزارش اکسل با موفقیت تولید شد")
        except Exception as e:
            logger.error(f"خطا در تولید گزارش اکسل: {e}")

        logger.info(f"تحلیل کامل شد. تعداد سیگنال‌های معتبر: {len(filtered_signals)}")
        
        if filtered_signals:
            symbols_text = ", ".join(f"{s['symbol']} {s['timeframe']}" for s in filtered_signals[:3])
            logger.info(f"سیگنال‌های برتر: {symbols_text}")
        else:
            logger.info("سیگنال قوی یافت نشد")

        if config.TELEGRAM_ENABLED and send_top_to_telegram and top_signals:
            message = format_top_signals(top_signals)
            try:
                await send_telegram_with_fallback(session, message)
            except Exception as e:
                logger.error(f"خطا در ارسال تلگرام: {e}")

    return top_signals

# --------- PERIODIC RUN LOOP ----------
async def run_periodically():
    logger.info("حالت اجرای دوره‌ای آغاز شد...")
    init_db()
    update_database_schema()
    while True:
        try:
            start = datetime.now(timezone.utc)  # ✅ اصلاح شد
            logger.info(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] شروع تحلیل...")
            top = await main_analysis(send_top_to_telegram=True)
            logger.info(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] تحلیل پایان یافت. تعداد برتر={len(top)}")
            elapsed = (datetime.now(timezone.utc) - start).total_seconds()  # ✅ اصلاح شد
            sleep_for = max(1, config.RUN_INTERVAL - elapsed)
            await asyncio.sleep(sleep_for)
        except KeyboardInterrupt:
            logger.info("حلقه متوقف شد.")
            break
        except Exception as e:
            logger.error(f"خطا در حلقه دوره‌ای: {e}", exc_info=True)
            await asyncio.sleep(60)

# --------- CLI ENTRY ----------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Leila Trading Bot - نسخه حرفه‌ای یکپارچه (فارسی)")
    parser.add_argument("--once", action="store_true", help="اجرای تحلیل یک‌باره و خروج")
    parser.add_argument("--loop", action="store_true", help="اجرای تحلیل دوره‌ای")
    parser.add_argument("--health", action="store_true", help="اجرای سرور سلامت/متریک‌ها")
    args = parser.parse_args()

    init_db()
    update_database_schema()
    logger.info("دیتابیس آماده است")

    if args.health:
        asyncio.run(start_health_server())

    if args.once:
        logger.info("اجرای تحلیل یک‌باره...")
        asyncio.run(main_analysis(send_top_to_telegram=True))
    elif args.loop:
        logger.info("اجرای حالت دوره‌ای...")
        asyncio.run(run_periodically())
    else:
        logger.info("اجرای پیش‌فرض (یک‌باره)...")
        asyncio.run(main_analysis(send_top_to_telegram=True))